from openpyxl import load_workbook
from openpyxl import Workbook
from os import remove

# Module Doc: https://openpyxl.readthedocs.io/en/stable/

class excel:
  @staticmethod
  def columnHeight(fileName, column):
    '''
    This module will handle importing the excel documents and populating the lists appropriately.
    '''
    wb = load_workbook(fileName)
    ws = wb.active
    rowNum = 1
    cell = str(column) + str(rowNum)
    while(ws[cell].value != None):
      rowNum+= 1
      cell = str(column) + str(rowNum)
    return rowNum

  @staticmethod
  def importColumn(fileName, column):
    '''
    This module will handle importing the excel documents and populating the lists appropriately.
    '''
    # Load workbook from file
    wb = load_workbook(fileName)
    ws = wb.active
    # create list of pages
    generatedList = []
    rowNum = 2
    cell = str(column) + str(rowNum)
    while(ws[cell].value != None):
      generatedList.append(ws[cell].value)
      rowNum+= 1
      cell = str(column) + str(rowNum)
    return generatedList

  @staticmethod
  def exportRow(ws, generatedList, row):
    '''
    This module will handle exporting the resulting list to excel
    '''
    i = 1
    for item in generatedList:
      ws.cell(row = row, column = i).value = item
      i += 1
    return
  @staticmethod
  def exportColumn(generatedList, col):
    '''
    This module will handle exporting the resulting list to excel
    '''
    i = 1
    wb = Workbook()
    ws = wb.active
    for item in generatedList:
      ws.cell(row = i, column = col).value = item
      i += 1
    return wb

  @staticmethod
  def exportFunc(currList, fileName):
    '''
    Wrapper for exportRow to conform to desired format
    '''
    if fileName in 'output.xlsx':
      row = excel.columnHeight(fileName,'a')
      wb = load_workbook(fileName)
      remove(fileName)
      finalWS = wb.active
      while currList:
        excel.exportRow(finalWS, currList.pop(0), row)
        row += 1
    else:
      remove(fileName)
      currList.insert(0,fileName.replace('.xlsx',''))
      wb = excel.exportColumn(currList, 1)
    wb.save(fileName)
    return
